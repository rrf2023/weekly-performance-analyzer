#!/usr/bin/env python3

import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import pandas as pd
import psycopg2
from psycopg2 import sql
from datetime import timedelta
from openpyxl.styles import Font


APP_W = 1024
APP_H = 640
SETUP_FILE = "setup.txt"
HH_LEN = 6          # household code length: 000629
CATEGORY_LEN = 3    # category code length: 011, 022, 098, 201, 831



class WeeklyPerformanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Weekly Performance Analyzer")
        self.root.geometry(f"{APP_W}x{APP_H}")
        self.conn = None

        self.build_gui()
        self.load_setup_file()
        self.root.protocol("WM_DELETE_WINDOW", self.exit_app)

    # ---------------- GUI ----------------
    def build_gui(self):
        frm = tk.Frame(self.root)
        frm.pack(padx=10, pady=10, fill="x")

        self.db_var = tk.StringVar()
        self.host_var = tk.StringVar()
        self.port_var = tk.StringVar(value="5432")
        self.user_var = tk.StringVar()
        self.pass_var = tk.StringVar()

        fields = [
            ("Database", self.db_var),
            ("Host", self.host_var),
            ("Port", self.port_var),
            ("User", self.user_var),
            ("Password", self.pass_var),
        ]

        for i, (lbl, var) in enumerate(fields):
            tk.Label(frm, text=lbl).grid(row=i, column=0, sticky="w")
            tk.Entry(
                frm,
                textvariable=var,
                width=40,
                show="*" if lbl == "Password" else ""
            ).grid(row=i, column=1, padx=5, pady=2)

        tk.Button(frm, text="Connect", command=self.connect_db).grid(
            row=0, column=2, rowspan=2, padx=10
        )

        row2 = tk.Frame(self.root)
        row2.pack(padx=10, pady=10, fill="x")

        tk.Label(row2, text="Table:").pack(side="left")
        self.table_cb = ttk.Combobox(row2, width=40, state="readonly")
        self.table_cb.pack(side="left", padx=5)
        self.table_cb.bind("<<ComboboxSelected>>", self.load_weeks)

        tk.Label(row2, text="Week Start:").pack(side="left", padx=(20, 0))
        self.week_cb = ttk.Combobox(row2, width=15, state="readonly")
        self.week_cb.pack(side="left", padx=5)

        tk.Button(row2, text="Run & Save Excel", command=self.run_all).pack(
            side="right", padx=10
        )
        tk.Button(row2, text="Exit", command=self.exit_app).pack(
            side="right", padx=10
        )

        self.log = tk.Text(self.root, height=20)
        self.log.pack(fill="both", expand=True, padx=10, pady=10)

    def write_log(self, txt):
        self.log.insert("end", txt + "\n")
        self.log.see("end")
        self.root.update_idletasks()

    # ---------------- Helpers ----------------
    @staticmethod
    def clean_text_value(x):
        if pd.isna(x):
            return ""
        return str(x).strip()

    @staticmethod
    def normalize_household(x):
        """
        Convert household code to zero-padded 6-digit string.
        Examples:
            629 -> 000629
            '1174.0' -> 001174
            '000629' -> 000629
        """
        if pd.isna(x):
            return ""

        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]

        digits = re.sub(r"\D", "", s)
        if not digits:
            return ""

        return digits.zfill(HH_LEN)

    @staticmethod
    def normalize_category(x):
        """
        Convert category code to zero-padded 3-digit string.
        Examples:
            11 -> 011
            '22.0' -> 022
            '071' -> 071
            '831' -> 831
        """
        if pd.isna(x):
            return ""

        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]

        digits = re.sub(r"\D", "", s)
        if not digits:
            return ""

        return digits.zfill(CATEGORY_LEN)

    @staticmethod
    def normalize_ean(x):
        """
        Convert EAN/SKU code to a clean digit string without losing leading zeroes.
        Examples:
            4870215280124 -> 4870215280124
            '4870215280124.0' -> 4870215280124
            ' 0460012345678 ' -> 0460012345678
        """
        if pd.isna(x):
            return ""

        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]

        digits = re.sub(r"\D", "", s)
        return digits

    @staticmethod
    def safe_filename(text):
        text = str(text).strip()
        if not text:
            text = "Без_интервьюера"
        text = re.sub(r'[\\/:*?"<>|]+', "_", text)
        text = re.sub(r"\s+", "_", text)
        return text[:120]

    @staticmethod
    def build_household_start_row_map(part):
        """
        Returns mapping:
            household -> first Excel row number in interviewer file
        Row numbers are for the output sheet where header is row 1,
        data starts at row 2.
        """
        hh_row_map = {}
        current_excel_row = 2  # row 1 = header

        for hh, hh_part in part.groupby("Домохозяйство", sort=False, dropna=False):
            hh = str(hh).strip()
            if hh and hh not in hh_row_map:
                hh_row_map[hh] = current_excel_row

            current_excel_row += len(hh_part)

        return hh_row_map

    @staticmethod
    def build_suspicious_summary_rows(suspicious_df, sku_df):
        """
        Build output rows for the summary workbook in the same visual structure
        as sample_1.xlsx:

        1) one household summary row
        2) underneath it, detail rows with EAN from f0122 and purchase counts

        Output columns:
            Домохозяйство, Интервьюер, Регион, Кол-во категорий,
            Признак_1_1или2_категории, Признак_2_копия_чека,
            Сигнатура, Кол-во покупок
        """
        out_rows = []

        if suspicious_df.empty:
            return pd.DataFrame(columns=[
                "Домохозяйство",
                "Интервьюер",
                "Регион",
                "Кол-во категорий",
                "Признак_1_1или2_категории",
                "Признак_2_копия_чека",
                "Сигнатура",
                "Кол-во покупок",
            ])

        sku_df = sku_df.copy()
        if not sku_df.empty:
            sku_df["EAN"] = sku_df["EAN"].astype(str).str.strip()
            sku_df["Количество покупок"] = pd.to_numeric(
                sku_df["Количество покупок"], errors="coerce"
            ).fillna(0).astype(int)

        suspicious_df = suspicious_df.sort_values(
            by=["Интервьюер", "Регион", "Домохозяйство"],
            ascending=[True, True, True]
        ).copy()

        for _, hh_row in suspicious_df.iterrows():
            hh = str(hh_row["Домохозяйство"]).strip()

            detail = sku_df[sku_df["Домохозяйство"].astype(str).str.strip() == hh].copy()
            if not detail.empty:
                detail = detail.sort_values(
                    by=["Количество покупок", "EAN"],
                    ascending=[False, True]
                )

            out_rows.append({
                "Домохозяйство": hh_row["Домохозяйство"],
                "Интервьюер": hh_row["Интервьюер"],
                "Регион": hh_row["Регион"],
                "Кол-во категорий": hh_row["Кол-во категорий"],
                "Признак_1_1или2_категории": hh_row["Признак_1_1или2_категории"],
                "Признак_2_копия_чека": hh_row["Признак_2_копия_чека"],
                "Сигнатура": hh_row["Сигнатура"],
                "Кол-во покупок": None,
            })

            for _, sku_row in detail.iterrows():
                out_rows.append({
                    "Домохозяйство": "",
                    "Интервьюер": "",
                    "Регион": "",
                    "Кол-во категорий": "",
                    "Признак_1_1или2_категории": "",
                    "Признак_2_копия_чека": "",
                    "Сигнатура": sku_row["EAN"],
                    "Кол-во покупок": int(sku_row["Количество покупок"]),
                })

        return pd.DataFrame(out_rows)

    @staticmethod
    def build_household_signature_map(category_part, sku_part):
        """
        Build suspicious HH summary for one interviewer.

        Rules:
        - Признак_1_1или2_категории: based on product groups/categories
        - Признак_2_копия_чека: based on exact EAN/SKU composition, not categories
        - Сигнатура shown in summary row: category-based, for human-readable overview

        category_part columns expected:
            Домохозяйство, Категория, Количество покупок, Интервьюер, Регион

        sku_part columns expected:
            Домохозяйство, EAN, Количество покупок, Интервьюер, Регион
        """
        if category_part.empty:
            return pd.DataFrame(columns=[
                "Домохозяйство",
                "Интервьюер",
                "Регион",
                "Кол-во категорий",
                "Признак_1_1или2_категории",
                "Признак_2_копия_чека",
                "Сигнатура"
            ])

        hh_rows = []
        ean_signature_to_hhs = {}

        category_part = category_part.copy()
        category_part["Категория"] = category_part["Категория"].astype(str).str.strip()
        category_part["Количество покупок"] = pd.to_numeric(
            category_part["Количество покупок"], errors="coerce"
        ).fillna(0).astype(int)

        sku_part = sku_part.copy()
        if not sku_part.empty:
            sku_part["Домохозяйство"] = sku_part["Домохозяйство"].astype(str).str.strip()
            sku_part["EAN"] = sku_part["EAN"].astype(str).str.strip()
            sku_part["Количество покупок"] = pd.to_numeric(
                sku_part["Количество покупок"], errors="coerce"
            ).fillna(0).astype(int)

        for hh, hh_part in category_part.groupby("Домохозяйство", dropna=False):
            hh_part = hh_part.copy()
            hh_str = str(hh).strip()

            interviewer = ""
            region = ""

            if "Интервьюер" in hh_part.columns and not hh_part["Интервьюер"].empty:
                interviewer = str(hh_part["Интервьюер"].iloc[0]).strip()

            if "Регион" in hh_part.columns and not hh_part["Регион"].empty:
                region = str(hh_part["Регион"].iloc[0]).strip()

            unique_categories = sorted(set(hh_part["Категория"].tolist()))
            cat_count = len(unique_categories)

            cat_agg = (
                hh_part.groupby("Категория", dropna=False)["Количество покупок"]
                .sum()
                .reset_index()
                .sort_values(by="Категория", ascending=True)
            )

            category_signature_items = [
                f"{row['Категория']}:{int(row['Количество покупок'])}"
                for _, row in cat_agg.iterrows()
                if str(row["Категория"]).strip()
            ]
            category_signature = " | ".join(category_signature_items)

            sku_hh = sku_part[
                sku_part["Домохозяйство"] == hh_str
            ].copy()

            ean_signature = ""
            if not sku_hh.empty:
                ean_agg = (
                    sku_hh.groupby("EAN", dropna=False)["Количество покупок"]
                    .sum()
                    .reset_index()
                    .sort_values(by="EAN", ascending=True)
                )

                ean_signature_items = [
                    f"{row['EAN']}:{int(row['Количество покупок'])}"
                    for _, row in ean_agg.iterrows()
                    if str(row["EAN"]).strip()
                ]
                ean_signature = " | ".join(ean_signature_items)

                if ean_signature.strip():
                    ean_signature_to_hhs.setdefault(ean_signature, set()).add(hh_str)

            hh_rows.append({
                "Домохозяйство": hh_str,
                "Интервьюер": interviewer,
                "Регион": region,
                "Кол-во категорий": cat_count,
                "Признак_1_1или2_категории": "Да" if cat_count in (1, 2) else "",
                "Признак_2_копия_чека": "",
                "Сигнатура": category_signature
            })

        duplicated_hhs = set()
        for ean_signature, hhs in ean_signature_to_hhs.items():
            if ean_signature.strip() and len(hhs) >= 2:
                duplicated_hhs.update(hhs)

        for row in hh_rows:
            if row["Домохозяйство"] in duplicated_hhs:
                row["Признак_2_копия_чека"] = "Да"

        out = pd.DataFrame(hh_rows)
        out = out[
            (out["Признак_1_1или2_категории"] == "Да") |
            (out["Признак_2_копия_чека"] == "Да")
        ].copy()

        out = out.sort_values(
            by=["Интервьюер", "Регион", "Домохозяйство"],
            ascending=[True, True, True]
        )
        return out

    # ---------------- Setup ----------------
    def load_setup_file(self):
        if not os.path.exists(SETUP_FILE):
            return

        cfg = {}
        with open(SETUP_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if "=" in line:
                    k, v = line.strip().split("=", 1)
                    cfg[k] = v

        self.host_var.set(cfg.get("host", ""))
        self.port_var.set(cfg.get("port", "5432"))
        self.db_var.set(cfg.get("database", ""))
        self.user_var.set(cfg.get("user", ""))
        self.pass_var.set(cfg.get("password", ""))

    def save_setup_file(self):
        with open(SETUP_FILE, "w", encoding="utf-8") as f:
            f.write(f"host={self.host_var.get()}\n")
            f.write(f"port={self.port_var.get()}\n")
            f.write(f"database={self.db_var.get()}\n")
            f.write(f"user={self.user_var.get()}\n")
            f.write(f"password={self.pass_var.get()}\n")

    def exit_app(self):
        self.save_setup_file()
        if self.conn:
            try:
                self.conn.close()
            except Exception:
                pass
        self.root.destroy()

    # ---------------- DB ----------------
    def connect_db(self):
        try:
            if self.conn:
                try:
                    self.conn.close()
                except Exception:
                    pass

            self.conn = psycopg2.connect(
                dbname=self.db_var.get().strip(),
                host=self.host_var.get().strip(),
                port=self.port_var.get().strip(),
                user=self.user_var.get().strip(),
                password=self.pass_var.get(),
            )

            self.write_log("Connected.")
            self.load_tables()

        except Exception as e:
            messagebox.showerror("Connection error", str(e))

    def get_valid_tables(self):
        q = """
            SELECT c.table_name
            FROM information_schema.columns c
            WHERE c.table_schema = 'public'
            GROUP BY c.table_name
            HAVING COUNT(DISTINCT CASE
                WHEN c.column_name IN ('f0103', 'f0105', 'f0122', 'prod_group', 'date')
                THEN c.column_name
            END) = 5
            ORDER BY c.table_name
        """
        return pd.read_sql(q, self.conn)

    def load_tables(self):
        df = self.get_valid_tables()
        tables = df["table_name"].tolist()
        self.table_cb["values"] = tables

        if tables:
            self.table_cb.current(0)
            self.load_weeks()
            self.write_log(f"Loaded {len(tables)} compatible table(s).")
        else:
            self.table_cb.set("")
            self.week_cb.set("")
            self.week_cb["values"] = []
            self.write_log(
                "No compatible tables found. Required columns: "
                "f0103, f0105, f0122, prod_group, date"
            )

    def validate_table_structure(self, table):
        q = """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
        """
        cols = pd.read_sql(q, self.conn, params=[table])["column_name"].tolist()
        required = {"f0103", "f0105", "f0122", "prod_group", "date"}
        missing = required - set(cols)

        if missing:
            raise Exception(
                f"Table '{table}' is missing required columns: {', '.join(sorted(missing))}"
            )

    def load_weeks(self, event=None):
        table = self.table_cb.get().strip()
        if not table:
            self.week_cb.set("")
            self.week_cb["values"] = []
            return

        try:
            self.validate_table_structure(table)

            q = sql.SQL("""
                SELECT DISTINCT date
                FROM {tbl}
                WHERE date IS NOT NULL
                ORDER BY date
            """).format(tbl=sql.Identifier(table))

            df = pd.read_sql(q.as_string(self.conn), self.conn)
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
            df = df.dropna(subset=["date"]).copy()

            if df.empty:
                self.week_cb.set("")
                self.week_cb["values"] = []
                return

            df["date"] = df["date"].dt.normalize()
            df["week_start"] = df["date"] - pd.to_timedelta(df["date"].dt.weekday, unit="d")

            weeks = sorted(
                df["week_start"].drop_duplicates().dt.strftime("%Y-%m-%d").tolist()
            )

            self.week_cb["values"] = weeks
            if weeks:
                self.week_cb.current(len(weeks) - 1)
            else:
                self.week_cb.set("")

        except Exception as e:
            self.week_cb.set("")
            self.week_cb["values"] = []
            self.write_log(f"Failed to load weeks for table '{table}': {e}")

    # ---------------- Excel reference files ----------------
    def load_supervisors(self):
        sup_path = os.path.join(os.getcwd(), "Supervisors.xlsx")
        if not os.path.exists(sup_path):
            raise Exception("Supervisors.xlsx not found in program folder")

        sup = pd.read_excel(sup_path)

        required_sup_cols = {"S2", "S4", "S5"}
        missing_sup = required_sup_cols - set(sup.columns)
        if missing_sup:
            raise Exception(
                f"Supervisors.xlsx is missing columns: {', '.join(sorted(missing_sup))}"
            )

        sup = sup.rename(columns={
            "S2": "hh_num",
            "S4": "Интервьюер",
            "S5": "Регион"
        })

        sup = sup[["hh_num", "Интервьюер", "Регион"]].copy()
        sup["hh_num"] = sup["hh_num"].apply(self.normalize_household)
        sup["Интервьюер"] = sup["Интервьюер"].apply(self.clean_text_value)
        sup["Регион"] = sup["Регион"].apply(self.clean_text_value)
        sup = sup[sup["hh_num"] != ""].drop_duplicates(subset=["hh_num"])

        return sup

    def load_reference(self):
        ref_path = os.path.join(os.getcwd(), "Справочник.xlsx")
        if not os.path.exists(ref_path):
            raise Exception("Справочник.xlsx not found in program folder")

        ref_df = pd.read_excel(ref_path)
        ref_df.columns = [str(c).strip() for c in ref_df.columns]

        required_ref_cols = {"Категория", "Описание"}
        missing_ref = required_ref_cols - set(ref_df.columns)
        if missing_ref:
            raise Exception(
                f"Справочник.xlsx is missing columns: {', '.join(sorted(missing_ref))}"
            )

        ref_df = ref_df[["Категория", "Описание"]].copy()
        ref_df["Категория"] = ref_df["Категория"].apply(self.normalize_category)
        ref_df["Описание"] = ref_df["Описание"].apply(self.clean_text_value)

        ref_df = ref_df[ref_df["Категория"] != ""].drop_duplicates(subset=["Категория"])
        ref_df = ref_df.rename(columns={"Категория": "ref_category"})

        return ref_df

    # ---------------- Main logic ----------------
    def run_all(self):
        table = self.table_cb.get().strip()
        week_start_str = self.week_cb.get().strip()

        if not table or not week_start_str:
            messagebox.showerror("Error", "Select table and week")
            return

        save_dir = filedialog.askdirectory(title="Select folder to save interviewer files")
        if not save_dir:
            return

        try:
            self.validate_table_structure(table)
            self.write_log("Loading reference files...")

            sup = self.load_supervisors()
            ref_df = self.load_reference()

            self.write_log("Loading database data...")

            data_q = sql.SQL("""
                SELECT
                    f0103 AS hh,
                    prod_group,
                    f0122 AS ean,
                    date
                FROM {tbl}
            """).format(tbl=sql.Identifier(table))

            df = pd.read_sql(data_q.as_string(self.conn), self.conn)

            if df.empty:
                raise Exception("Selected table contains no data")

            df["hh"] = df["hh"].apply(self.normalize_household)
            df["prod_group"] = df["prod_group"].apply(self.normalize_category)
            df["ean"] = df["ean"].apply(self.normalize_ean)
            df["date"] = pd.to_datetime(df["date"], errors="coerce")

            df = df[
                (df["hh"] != "") &
                (df["prod_group"] != "") &
                (df["date"].notna())
            ].copy()

            if df.empty:
                raise Exception("No valid rows found after cleaning source data")

            df["date"] = df["date"].dt.normalize()

            week_start = pd.to_datetime(week_start_str).normalize()
            week_end = week_start + timedelta(days=6)

            df_week = df[
                (df["date"] >= week_start) &
                (df["date"] <= week_end)
            ].copy()

            if df_week.empty:
                raise Exception(
                    f"No data found for selected week: "
                    f"{week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')}"
                )

            self.write_log(f"Rows in selected week: {len(df_week)}")

            result = (
                df_week.groupby(["hh", "prod_group"], dropna=False)
                .size()
                .reset_index(name="Количество покупок")
            )

            sku_result = (
                df_week[df_week["ean"] != ""].groupby(["hh", "ean"], dropna=False)
                .size()
                .reset_index(name="Количество покупок")
            )

            result = result.merge(
                sup,
                left_on="hh",
                right_on="hh_num",
                how="left"
            )

            result = result.merge(
                ref_df,
                left_on="prod_group",
                right_on="ref_category",
                how="left"
            )

            for col in ["hh_num", "ref_category"]:
                if col in result.columns:
                    result = result.drop(columns=[col])

            missing_desc = result[
                result["Описание"].isna() |
                (result["Описание"].astype(str).str.strip() == "")
            ]
            if not missing_desc.empty:
                missing_codes = sorted(missing_desc["prod_group"].astype(str).unique())
                self.write_log(
                    "Categories missing in reference: " + ", ".join(missing_codes)
                )

            result = result.rename(columns={
                "hh": "Домохозяйство",
                "prod_group": "Категория"
            })

            sku_result = sku_result.merge(
                sup,
                left_on="hh",
                right_on="hh_num",
                how="left"
            )
            for col in ["hh_num"]:
                if col in sku_result.columns:
                    sku_result = sku_result.drop(columns=[col])

            sku_result = sku_result.rename(columns={
                "hh": "Домохозяйство",
                "ean": "EAN"
            })

            result["Интервьюер"] = result["Интервьюер"].fillna("").astype(str).str.strip()
            result["Регион"] = result["Регион"].fillna("").astype(str).str.strip()
            result["Описание"] = result["Описание"].fillna("").astype(str).str.strip()

            sku_result["Интервьюер"] = sku_result["Интервьюер"].fillna("").astype(str).str.strip()
            sku_result["Регион"] = sku_result["Регион"].fillna("").astype(str).str.strip()

            result = result[
                [
                    "Домохозяйство",
                    "Категория",
                    "Описание",
                    "Количество покупок",
                    "Интервьюер",
                    "Регион"
                ]
            ].sort_values(
                by=["Интервьюер", "Домохозяйство", "Количество покупок", "Категория"],
                ascending=[True, True, False, True]
            )

            saved_count = 0
            suspicious_parts = []
            hh_link_map = {}

            for interviewer, part in result.groupby("Интервьюер", dropna=False):
                interviewer_name = interviewer.strip() if isinstance(interviewer, str) else ""
                if not interviewer_name:
                    interviewer_name = "Без_интервьюера"

                filename = (
                    f"{self.safe_filename(interviewer_name)}_"
                    f"{week_start.strftime('%Y-%m-%d')}.xlsx"
                )
                save_path = os.path.join(save_dir, filename)

                part = part.sort_values(
                    by=["Домохозяйство", "Количество покупок", "Категория"],
                    ascending=[True, False, True]
                ).copy()

                hh_start_rows = self.build_household_start_row_map(part)

                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    part.to_excel(writer, sheet_name="Weekly Performance", index=False)

                    ws = writer.sheets["Weekly Performance"]
                    widths = {
                        "A": 16,  # Домохозяйство
                        "B": 14,  # Категория
                        "C": 55,  # Описание
                        "D": 20,  # Количество покупок
                        "E": 22,  # Интервьюер
                        "F": 16,  # Регион
                    }
                    for col, width in widths.items():
                        ws.column_dimensions[col].width = width

                for hh, start_row in hh_start_rows.items():
                    hh_link_map[(interviewer_name, hh)] = {
                        "file_name": filename,
                        "sheet_name": "Weekly Performance",
                        "cell": f"A{start_row}"
                    }

                sku_part = sku_result[
                    sku_result["Интервьюер"].fillna("").astype(str).str.strip() == interviewer_name
                ].copy()

                suspicious_part = self.build_household_signature_map(part, sku_part)
                if not suspicious_part.empty:
                    suspicious_parts.append(suspicious_part)
                    self.write_log(
                        f"{interviewer_name}: suspicious HH found: {len(suspicious_part)}"
                    )

                self.write_log(f"Saved: {save_path}")
                saved_count += 1

            suspicious_save_path = os.path.join(
                save_dir,
                f"Suspicious_households_{week_start.strftime('%Y-%m-%d')}.xlsx"
            )

            if suspicious_parts:
                suspicious_all = pd.concat(suspicious_parts, ignore_index=True)

                suspicious_summary_rows = self.build_suspicious_summary_rows(
                    suspicious_df=suspicious_all,
                    sku_df=sku_result[["Домохозяйство", "EAN", "Количество покупок"]].copy()
                )

                with pd.ExcelWriter(suspicious_save_path, engine="openpyxl") as writer:
                    suspicious_summary_rows.to_excel(
                        writer,
                        sheet_name="Suspicious HH",
                        index=False
                    )

                    ws = writer.sheets["Suspicious HH"]
                    widths = {
                        "A": 16,  # Домохозяйство
                        "B": 22,  # Интервьюер
                        "C": 16,  # Регион
                        "D": 18,  # Кол-во категорий
                        "E": 22,  # Признак 1
                        "F": 20,  # Признак 2
                        "G": 80,  # Сигнатура / EAN
                        "H": 18,  # Кол-во покупок
                    }
                    for col, width in widths.items():
                        ws.column_dimensions[col].width = width

                    # Put total purchase formula into household rows, summing the detail EAN rows below.
                    # Also group EAN detail rows under each HH row, so the user can collapse/expand them.
                    ws.sheet_properties.outlinePr.summaryBelow = True
                    ws.freeze_panes = "A2"

                    current_row = 2
                    max_row = ws.max_row
                    while current_row <= max_row:
                        hh_value = ws[f"A{current_row}"].value
                        if hh_value not in (None, ""):
                            # make household row visually stand out
                            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                                ws[f"{col_letter}{current_row}"].font = Font(bold=True)

                            hh = str(ws[f"A{current_row}"].value).strip()
                            interviewer_name = str(ws[f"B{current_row}"].value).strip()
                            link_info = hh_link_map.get((interviewer_name, hh))
                            if link_info:
                                file_name = str(link_info["file_name"]).replace("'", "''")
                                sheet_name = str(link_info["sheet_name"]).replace("'", "''")
                                cell_ref = link_info["cell"]
                                target = f"{file_name}#'{sheet_name}'!{cell_ref}"

                                for link_col in ["A", "G"]:
                                    cell = ws[f"{link_col}{current_row}"]
                                    cell.hyperlink = target
                                    cell.font = Font(bold=True, color="0000FF", underline="single")

                            start_detail = current_row + 1
                            end_detail = start_detail - 1

                            while end_detail + 1 <= max_row and ws[f"A{end_detail + 1}"].value in (None, ""):
                                end_detail += 1

                            if end_detail >= start_detail:
                                ws[f"H{current_row}"] = f"=SUM(H{start_detail}:H{end_detail})"

                                # group detail rows under the HH summary row
                                for detail_row in range(start_detail, end_detail + 1):
                                    dim = ws.row_dimensions[detail_row]
                                    dim.outlineLevel = 1
                                    dim.hidden = True
                            else:
                                ws[f"H{current_row}"] = 0

                            current_row = end_detail + 1
                        else:
                            current_row += 1

                self.write_log(f"Saved suspicious summary: {suspicious_save_path}")
            else:
                self.write_log("No suspicious households found for summary file.")

            self.write_log(f"DONE: created {saved_count} interviewer file(s) in {save_dir}")
            messagebox.showinfo("Done", f"Created {saved_count} interviewer file(s).")

        except Exception as e:
            self.write_log(f"ERROR: {e}")
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = WeeklyPerformanceApp(root)
    root.mainloop()