#!/usr/bin/env python3

import os
import re

import pandas as pd
from openpyxl.styles import Font


class ExcelExporter:
    # ==========================
    # Helpers
    # ==========================
    @staticmethod
    def safe_filename(text):
        text = str(text).strip()

        if not text:
            text = "Без_интервьюера"

        text = re.sub(
            r'[\\/:*?"<>|]+',
            "_",
            text
        )

        text = re.sub(
            r"\s+",
            "_",
            text
        )

        return text[:120]

    @staticmethod
    def build_household_start_row_map(part):
        """
        Returns mapping:

            household -> first Excel row

        Header = row 1
        Data starts = row 2
        """
        household_map = {}
        current_excel_row = 2

        for hh, hh_part in part.groupby(
            "Домохозяйство",
            sort=False,
            dropna=False
        ):
            hh = str(hh).strip()

            if hh and hh not in household_map:
                household_map[hh] = (
                    current_excel_row
                )

            current_excel_row += len(
                hh_part
            )

        return household_map

    @staticmethod
    def build_suspicious_summary_rows(
        suspicious_df,
        sku_df
    ):
        """
        Build summary workbook rows.

        Structure:
        1. HH summary row
        2. Detail rows below
           with EAN and count
        """

        output_rows = []

        if suspicious_df.empty:
            return pd.DataFrame(columns=[
                "Домохозяйство",
                "Интервьюер",
                "Регион",
                "Кол-во категорий",
                "Признак_1_1или2_категории",
                "Признак_2_копия_чека",
                "Сигнатура",
                "Кол-во покупок",
            ])

        sku_df = sku_df.copy()

        sku_df["EAN"] = (
            sku_df["EAN"]
            .astype(str)
            .str.strip()
        )

        sku_df["Количество покупок"] = (
            pd.to_numeric(
                sku_df["Количество покупок"],
                errors="coerce"
            )
            .fillna(0)
            .astype(int)
        )

        suspicious_df = (
            suspicious_df
            .sort_values(
                by=[
                    "Интервьюер",
                    "Регион",
                    "Домохозяйство"
                ]
            )
            .copy()
        )

        for _, hh_row in (
            suspicious_df.iterrows()
        ):
            hh = str(
                hh_row["Домохозяйство"]
            ).strip()

            detail = sku_df[
                sku_df[
                    "Домохозяйство"
                ]
                .astype(str)
                .str.strip()
                == hh
            ].copy()

            if not detail.empty:
                detail = detail.sort_values(
                    by=[
                        "Количество покупок",
                        "EAN"
                    ],
                    ascending=[
                        False,
                        True
                    ]
                )

            output_rows.append({
                "Домохозяйство":
                    hh_row[
                        "Домохозяйство"
                    ],
                "Интервьюер":
                    hh_row[
                        "Интервьюер"
                    ],
                "Регион":
                    hh_row["Регион"],
                "Кол-во категорий":
                    hh_row[
                        "Кол-во категорий"
                    ],
                "Признак_1_1или2_категории":
                    hh_row[
                        "Признак_1_1или2_категории"
                    ],
                "Признак_2_копия_чека":
                    hh_row[
                        "Признак_2_копия_чека"
                    ],
                "Сигнатура":
                    hh_row[
                        "Сигнатура"
                    ],
                "Кол-во покупок":
                    None
            })

            for _, sku_row in (
                detail.iterrows()
            ):
                output_rows.append({
                    "Домохозяйство":
                        "",
                    "Интервьюер":
                        "",
                    "Регион":
                        "",
                    "Кол-во категорий":
                        "",
                    "Признак_1_1или2_категории":
                        "",
                    "Признак_2_копия_чека":
                        "",
                    "Сигнатура":
                        sku_row["EAN"],
                    "Кол-во покупок":
                        int(
                            sku_row[
                                "Количество покупок"
                            ]
                        )
                })

        return pd.DataFrame(
            output_rows
        )

    # ==========================
    # Main export
    # ==========================
    def export(
        self,
        result_bundle,
        save_dir,
        week_start,
        logger=None
    ):
        result = (
            result_bundle["result"]
            .copy()
        )

        sku_result = (
            result_bundle["sku_result"]
            .copy()
        )

        suspicious_df = (
            result_bundle["suspicious"]
            .copy()
        )

        week_str = str(
            pd.to_datetime(
                week_start
            ).strftime("%Y-%m-%d")
        )

        saved_count = 0
        household_link_map = {}

        # ==========================
        # Interviewer files
        # ==========================
        for interviewer, part in (
            result.groupby(
                "Интервьюер",
                dropna=False
            )
        ):
            interviewer_name = (
                str(interviewer)
                .strip()
            )

            if not interviewer_name:
                interviewer_name = (
                    "Без_интервьюера"
                )

            filename = (
                f"{self.safe_filename(interviewer_name)}_"
                f"{week_str}.xlsx"
            )

            save_path = os.path.join(
                save_dir,
                filename
            )

            part = (
                part.sort_values(
                    by=[
                        "Домохозяйство",
                        "Количество покупок",
                        "Категория"
                    ],
                    ascending=[
                        True,
                        False,
                        True
                    ]
                )
                .copy()
            )

            household_rows = (
                self
                .build_household_start_row_map(
                    part
                )
            )

            with pd.ExcelWriter(
                save_path,
                engine="openpyxl"
            ) as writer:

                part.to_excel(
                    writer,
                    sheet_name="Weekly Performance",
                    index=False
                )

                ws = writer.sheets[
                    "Weekly Performance"
                ]

                widths = {
                    "A": 16,
                    "B": 14,
                    "C": 55,
                    "D": 20,
                    "E": 22,
                    "F": 16,
                }

                for col, width in (
                    widths.items()
                ):
                    ws.column_dimensions[
                        col
                    ].width = width

            for hh, row_num in (
                household_rows.items()
            ):
                household_link_map[
                    (
                        interviewer_name,
                        hh
                    )
                ] = {
                    "file_name":
                        filename,
                    "sheet_name":
                        "Weekly Performance",
                    "cell":
                        f"A{row_num}"
                }

            if logger:
                logger(
                    f"Saved: "
                    f"{save_path}"
                )

            saved_count += 1

        # ==========================
        # Suspicious summary
        # ==========================
        suspicious_save_path = (
            os.path.join(
                save_dir,
                f"Suspicious_households_"
                f"{week_str}.xlsx"
            )
        )

        if not suspicious_df.empty:
            summary_rows = (
                self
                .build_suspicious_summary_rows(
                    suspicious_df=
                        suspicious_df,
                    sku_df=
                        sku_result[
                            [
                                "Домохозяйство",
                                "EAN",
                                "Количество покупок"
                            ]
                        ].copy()
                )
            )

            with pd.ExcelWriter(
                suspicious_save_path,
                engine="openpyxl"
            ) as writer:

                summary_rows.to_excel(
                    writer,
                    sheet_name="Suspicious HH",
                    index=False
                )

                ws = writer.sheets[
                    "Suspicious HH"
                ]

                widths = {
                    "A": 16,
                    "B": 22,
                    "C": 16,
                    "D": 18,
                    "E": 22,
                    "F": 20,
                    "G": 80,
                    "H": 18,
                }

                for col, width in (
                    widths.items()
                ):
                    ws.column_dimensions[
                        col
                    ].width = width

                ws.sheet_properties.outlinePr.summaryBelow = True
                ws.freeze_panes = "A2"

                current_row = 2
                max_row = ws.max_row

                while current_row <= max_row:
                    hh_value = ws[
                        f"A{current_row}"
                    ].value

                    if hh_value not in (
                        None,
                        ""
                    ):
                        for col_letter in [
                            "A",
                            "B",
                            "C",
                            "D",
                            "E",
                            "F",
                            "G",
                            "H"
                        ]:
                            ws[
                                f"{col_letter}"
                                f"{current_row}"
                            ].font = Font(
                                bold=True
                            )

                        hh = str(
                            ws[
                                f"A{current_row}"
                            ].value
                        ).strip()

                        interviewer_name = str(
                            ws[
                                f"B{current_row}"
                            ].value
                        ).strip()

                        link_info = (
                            household_link_map.get(
                                (
                                    interviewer_name,
                                    hh
                                )
                            )
                        )

                        if link_info:
                            file_name = (
                                str(
                                    link_info[
                                        "file_name"
                                    ]
                                )
                                .replace(
                                    "'",
                                    "''"
                                )
                            )

                            sheet_name = (
                                str(
                                    link_info[
                                        "sheet_name"
                                    ]
                                )
                                .replace(
                                    "'",
                                    "''"
                                )
                            )

                            target = (
                                f"{file_name}"
                                f"#'{sheet_name}'!"
                                f"{link_info['cell']}"
                            )

                            for col in [
                                "A",
                                "G"
                            ]:
                                cell = ws[
                                    f"{col}"
                                    f"{current_row}"
                                ]

                                cell.hyperlink = (
                                    target
                                )

                                cell.font = Font(
                                    bold=True,
                                    color="0000FF",
                                    underline="single"
                                )

                        start_detail = (
                            current_row + 1
                        )

                        end_detail = (
                            start_detail - 1
                        )

                        while (
                            end_detail + 1
                            <= max_row
                            and ws[
                                f"A{end_detail + 1}"
                            ].value
                            in (
                                None,
                                ""
                            )
                        ):
                            end_detail += 1

                        if (
                            end_detail
                            >= start_detail
                        ):
                            ws[
                                f"H{current_row}"
                            ] = (
                                f"=SUM("
                                f"H{start_detail}:"
                                f"H{end_detail}"
                                f")"
                            )

                            for detail_row in range(
                                start_detail,
                                end_detail + 1
                            ):
                                row_dim = (
                                    ws.row_dimensions[
                                        detail_row
                                    ]
                                )

                                row_dim.outlineLevel = 1
                                row_dim.hidden = True
                        else:
                            ws[
                                f"H{current_row}"
                            ] = 0

                        current_row = (
                            end_detail + 1
                        )

                    else:
                        current_row += 1

            if logger:
                logger(
                    "Saved suspicious summary: "
                    f"{suspicious_save_path}"
                )

        else:
            if logger:
                logger(
                    "No suspicious households "
                    "found for summary file."
                )

        return saved_count
